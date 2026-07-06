"""
analyzer.py
===========
Módulo 3 — Motor de Inteligencia y Filtro Comercial.

Lee los .xlsx descargados por el Módulo 2 (uno por marca), cruza cada fila
contra el Registro de Importadores y las Marcas Competidoras de Notion, y
genera:

  - reporte_marcas_mensual.xlsx  -> reporte consolidado con estilo corporativo
  - dict de resumen (para el Módulo 4 / WhatsApp) con los contadores:
        alertas_fuga, alertas_precio, nuevos_hallazgos

Reglas de negocio:
  1. Importador NO_COMPITE           -> se descarta la fila (ruido de otros rubros).
  2. Importador inexistente en Notion -> se guarda en NUEVOS_HALLAZGOS y se
                                         da de alta en Notion como POR_VALIDAR.
  3. Importador == CLIENTE y compra marca rival -> "Alerta Urgente de Pérdida
                                                    de Cliente".
  4. Importador == COMPETIDOR y precio unitario aduana < Precio_Techo_Nuestro
     * (1 - UMBRAL_PRESION_PRECIO) -> "Alerta de Competencia por Precio Bajo".

Marcas propias protegidas (nunca se tratan como rivales): COEL, SANHUA, VHM.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from notion_client import (
    get_marcas_competidoras,
    get_registro_importadores,
    alta_importador_por_validar,
)

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

# ------------------------------------------------------------------
# Configuración de negocio
# ------------------------------------------------------------------

MARCAS_PROPIAS_PROTEGIDAS = {"COEL", "SANHUA", "VHM", "K11"}

# % configurable: cuánto más bajo debe ser el precio de la competencia
# respecto a nuestro Precio_Techo_Nuestro para disparar la alerta.
UMBRAL_PRESION_PRECIO = 0.20  # 20%

ESTADO_NO_COMPITE = "NO_COMPITE"
ESTADO_CLIENTE = "CLIENTE"
ESTADO_COMPETIDOR = "COMPETIDOR"
ESTADO_POR_VALIDAR = "POR_VALIDAR"

# Nombres de columnas esperadas en los .xlsx exportados por Penta.
# AJUSTAR si el export real usa otros encabezados.
COL_IMPORTADOR = "Importador"
COL_MARCA = "Marca"
COL_PRECIO_UNITARIO = "Precio_Unitario_FOB"
COL_PRODUCTO = "Producto"
COL_CANTIDAD = "Cantidad"
COL_FECHA = "Fecha"

REPORTE_SALIDA = Path("reporte_marcas_mensual.xlsx")


@dataclass
class ResultadoAnalisis:
    filas_validas: List[dict] = field(default_factory=list)
    alertas_fuga: List[dict] = field(default_factory=list)
    alertas_precio: List[dict] = field(default_factory=list)
    nuevos_hallazgos: List[str] = field(default_factory=list)


def _normalizar_importador(valor: str) -> str:
    return " ".join(str(valor).split()).upper()


def _cargar_excels_descargados(carpeta: Path = Path("downloads")) -> pd.DataFrame:
    """Concatena todos los datos_{marca}.xlsx descargados en un único DataFrame."""
    archivos = sorted(carpeta.glob("datos_*.xlsx"))
    if not archivos:
        logger.warning("No se encontraron archivos datos_*.xlsx en %s", carpeta)
        return pd.DataFrame()

    dataframes = []
    for archivo in archivos:
        try:
            df = pd.read_excel(archivo)
            df["__archivo_origen"] = archivo.name
            dataframes.append(df)
        except Exception:
            logger.exception("No se pudo leer %s", archivo)

    if not dataframes:
        return pd.DataFrame()

    return pd.concat(dataframes, ignore_index=True)


def analizar(df: pd.DataFrame | None = None) -> ResultadoAnalisis:
    """
    Ejecuta la lógica de análisis fila por fila descripta en el módulo.
    Si no se pasa `df`, carga automáticamente los datos_*.xlsx de downloads/.
    """
    if df is None:
        df = _cargar_excels_descargados()

    resultado = ResultadoAnalisis()

    if df.empty:
        logger.warning("DataFrame vacío; no hay nada que analizar.")
        return resultado

    marcas_rivales, precios_techo, marca_propia_map = get_marcas_competidoras()
    marcas_rivales_set = {m.upper() for m in marcas_rivales}
    registro_importadores = get_registro_importadores()

    # Cache local para no dar de alta el mismo importador nuevo más de una vez
    # por corrida.
    importadores_nuevos_vistos: set[str] = set()

    for _, fila in df.iterrows():
        importador_raw = fila.get(COL_IMPORTADOR, "")
        if pd.isna(importador_raw) or not str(importador_raw).strip():
            continue

        importador = _normalizar_importador(importador_raw)
        marca_raw = fila.get(COL_MARCA, "")
        marca = str(marca_raw).strip().upper() if pd.notna(marca_raw) else ""

        # Nunca tratamos nuestras propias marcas como "rivales".
        if marca in MARCAS_PROPIAS_PROTEGIDAS:
            continue

        estado = registro_importadores.get(importador)

        # Regla 1: descartar ruido de importadores que no compiten.
        if estado == ESTADO_NO_COMPITE:
            continue

        # Regla 2: importador nuevo -> alta automática como POR_VALIDAR.
        if estado is None:
            if importador not in importadores_nuevos_vistos:
                resultado.nuevos_hallazgos.append(importador)
                importadores_nuevos_vistos.add(importador)
                try:
                    alta_importador_por_validar(importador)
                except Exception:
                    logger.exception("No se pudo dar de alta '%s' en Notion", importador)
            estado = ESTADO_POR_VALIDAR

        fila_dict = fila.to_dict()
        fila_dict[COL_IMPORTADOR] = importador
        fila_dict[COL_MARCA] = marca
        fila_dict["Estado_Importador"] = estado
        # A cuál de nuestras marcas propias (COEL/SANHUA/VHM) le compite esta fila.
        fila_dict["Marca_Propia_Afectada"] = marca_propia_map.get(marca, "SIN_ASIGNAR")

        precio_unitario = fila.get(COL_PRECIO_UNITARIO)
        try:
            precio_unitario = float(precio_unitario)
        except (TypeError, ValueError):
            precio_unitario = None
        fila_dict[COL_PRECIO_UNITARIO] = precio_unitario

        # Regla 3: fuga de cliente propio hacia marca rival.
        if estado == ESTADO_CLIENTE and marca in marcas_rivales_set:
            alerta = {
                **fila_dict,
                "Tipo_Alerta": "🚨 Alerta Urgente de Pérdida de Cliente",
            }
            resultado.alertas_fuga.append(alerta)

        # Regla 4: presión de precio de un competidor.
        if estado == ESTADO_COMPETIDOR and marca in marcas_rivales_set and precio_unitario:
            precio_techo = precios_techo.get(marca)
            if precio_techo:
                umbral = precio_techo * (1 - UMBRAL_PRESION_PRECIO)
                if precio_unitario < umbral:
                    alerta = {
                        **fila_dict,
                        "Precio_Techo_Nuestro": precio_techo,
                        "Umbral_Alerta": round(umbral, 2),
                        "Tipo_Alerta": "📉 Alerta de Competencia por Precio Bajo",
                    }
                    resultado.alertas_precio.append(alerta)

        resultado.filas_validas.append(fila_dict)

    logger.info(
        "Análisis finalizado: %d filas válidas, %d fugas, %d alertas de precio, %d nuevos hallazgos",
        len(resultado.filas_validas), len(resultado.alertas_fuga),
        len(resultado.alertas_precio), len(resultado.nuevos_hallazgos),
    )
    return resultado


# ------------------------------------------------------------------
# Generación del reporte Excel con estilo corporativo
# ------------------------------------------------------------------

AZUL_MARINO = "1F3864"
BLANCO = "FFFFFF"
GRIS_CEBRA = "F2F2F2"

HEADER_FILL = PatternFill(start_color=AZUL_MARINO, end_color=AZUL_MARINO, fill_type="solid")
HEADER_FONT = Font(color=BLANCO, bold=True, size=11)
ZEBRA_FILL = PatternFill(start_color=GRIS_CEBRA, end_color=GRIS_CEBRA, fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

COLUMNAS_MONEDA = {COL_PRECIO_UNITARIO, "Precio_Techo_Nuestro", "Umbral_Alerta"}


def _escribir_hoja(wb: Workbook, nombre_hoja: str, registros: List[dict]) -> None:
    ws = wb.create_sheet(title=nombre_hoja[:31])

    if not registros:
        ws.append(["Sin datos para esta sección."])
        ws["A1"].font = Font(italic=True, color="808080")
        return

    columnas = list({k for r in registros for k in r.keys()})
    # Orden estable: Importador, Marca, Producto primero si existen.
    prioridad = [COL_IMPORTADOR, COL_MARCA, COL_PRODUCTO, "Tipo_Alerta"]
    columnas_ordenadas = [c for c in prioridad if c in columnas] + [
        c for c in columnas if c not in prioridad and not c.startswith("__")
    ]

    ws.append(columnas_ordenadas)
    for col_idx, _ in enumerate(columnas_ordenadas, start=1):
        celda = ws.cell(row=1, column=col_idx)
        celda.fill = HEADER_FILL
        celda.font = HEADER_FONT
        celda.alignment = Alignment(horizontal="center", vertical="center")
        celda.border = THIN_BORDER

    for row_idx, registro in enumerate(registros, start=2):
        for col_idx, col in enumerate(columnas_ordenadas, start=1):
            valor = registro.get(col, "")
            celda = ws.cell(row=row_idx, column=col_idx, value=valor)
            celda.border = THIN_BORDER

            if col in COLUMNAS_MONEDA and isinstance(valor, (int, float)):
                celda.number_format = '"$"#,##0.00'

            if row_idx % 2 == 0:
                celda.fill = ZEBRA_FILL

    for col_idx, col in enumerate(columnas_ordenadas, start=1):
        max_len = max(
            [len(str(col))] + [len(str(r.get(col, ""))) for r in registros]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

    ws.freeze_panes = "A2"


def generar_reporte_excel(resultado: ResultadoAnalisis, salida: Path = REPORTE_SALIDA) -> Path:
    wb = Workbook()
    wb.remove(wb.active)  # quitar hoja default

    _escribir_hoja(wb, "Alertas Fuga Clientes", resultado.alertas_fuga)
    _escribir_hoja(wb, "Alertas Precio Bajo", resultado.alertas_precio)
    _escribir_hoja(wb, "Nuevos Hallazgos", [{"Importador": i} for i in resultado.nuevos_hallazgos])
    _escribir_hoja(wb, "Detalle Completo", resultado.filas_validas)

    wb.save(salida)
    logger.info("Reporte generado: %s", salida)
    return salida


def _desglose_por_marca_propia(resultado: ResultadoAnalisis) -> Dict[str, dict]:
    """Agrupa los conteos de alertas por marca propia (COEL/SANHUA/VHM/...)."""
    desglose: Dict[str, dict] = {}

    def _sumar(lista: List[dict], clave: str) -> None:
        for item in lista:
            marca_propia = item.get("Marca_Propia_Afectada", "SIN_ASIGNAR")
            desglose.setdefault(marca_propia, {"alertas_fuga": 0, "alertas_precio": 0})
            desglose[marca_propia][clave] += 1

    _sumar(resultado.alertas_fuga, "alertas_fuga")
    _sumar(resultado.alertas_precio, "alertas_precio")
    return desglose


def run() -> dict:
    """Punto de entrada del módulo: analiza y genera el reporte. Devuelve
    el resumen numérico que consume el Módulo 4 (WhatsApp)."""
    resultado = analizar()
    generar_reporte_excel(resultado)

    return {
        "alertas_fuga": len(resultado.alertas_fuga),
        "alertas_precio": len(resultado.alertas_precio),
        "nuevos_hallazgos": len(resultado.nuevos_hallazgos),
        "reporte_path": str(REPORTE_SALIDA),
        "desglose_por_marca_propia": _desglose_por_marca_propia(resultado),
    }


if __name__ == "__main__":
    print(run())
